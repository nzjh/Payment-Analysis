"""
Created on Sat Jun 27 10:05:08 2020
Author: Jun Zhao
Python Version: 3

This script reads in an Excel file and calculate the percentage of each category.
"""


import pandas as pd
from openpyxl import load_workbook

def execute(name):
    df_in = pd.read_excel('2020.xlsx', sheet_name=name)
    category_dict = dict()
    total = 0;
    
    
    #Calculate total payment for each category
    for index in df_in.index:
        category = df_in.loc[df_in.index[index], 'Category']
        price = df_in.loc[df_in.index[index], 'Price']
        total += price
        if(category not in category_dict):
            category_dict[category] = price
        else:
            category_dict[category] = category_dict[category] + price

                        
    #Construct output dataframe
    payments = category_dict.values()
    payments = sorted(payments, reverse=True)
    categories = list()
    percentages = list()
    for payment in payments:
        category = list(category_dict.keys())[list(category_dict.values()).index(payment)]
        categories.append(category)
        percentage = payment/total*100
        percentages.append(percentage)
    categories.append('Total')
    payments.append(total)
    percentages.append(100)
    df_out = pd.DataFrame({'Category':categories,
                           'Payment':payments,
                           'Percentage':percentages})
    
    
    #Write to source file
    book = load_workbook('2020.xlsx')
    writer = pd.ExcelWriter('2020.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df_out.to_excel(writer, sheet_name=name, startcol=4)
    writer.save()

wb = load_workbook('2020.xlsx')
sheet_names = wb.get_sheet_names()
for name in sheet_names:
    execute(name)
